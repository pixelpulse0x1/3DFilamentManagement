"""Base routes: pages, settings, background management, utilities."""
import os
import re
import io
import zipfile
import logging
from datetime import datetime
from openpyxl import Workbook
from flask import current_app, render_template, jsonify, request, send_file, send_from_directory

from modules.base import base_bp
from modules.db import get_db
from modules.db import LATEST_VERSION
from modules.base.bg_utils import (
    get_active_background, list_backgrounds, upload_background, set_active_background, get_background_dir,
)
from modules.base.migrate_utils import migrate_from_db, migrate_from_txt

logger = logging.getLogger(__name__)


def _data_dir():
    return current_app.config.get("DATA_DIR", os.path.join(os.path.dirname(__file__), "..", "..", "data"))


def _get_appearance_settings(conn):
    """Load card_opacity and card_color from system_settings with validation fallback."""
    rows = conn.execute(
        "SELECT key, value FROM system_settings WHERE key IN ('card_opacity', 'card_color', 'card_blur')"
    ).fetchall()
    sys = {r["key"]: r["value"] for r in rows}

    try:
        card_opacity = float(sys.get("card_opacity", "0.15"))
        if not (0.0 <= card_opacity <= 1.0):
            card_opacity = 0.15
    except (ValueError, TypeError):
        card_opacity = 0.15

    try:
        card_color = sys.get("card_color", "#ffffff")
        if not re.match(r'^#[0-9a-fA-F]{6}$', card_color):
            card_color = "#ffffff"
    except (ValueError, TypeError):
        card_color = "#ffffff"

    try:
        card_blur = int(sys.get("card_blur", "1"))
        if not (0 <= card_blur <= 30):
            card_blur = 1
    except (ValueError, TypeError):
        card_blur = 1

    return card_opacity, card_color, card_blur


def _bg_for_template():
    """Helper to get background for page renders."""
    return get_active_background(_data_dir())


# ─── Page Routes ───

@base_bp.route("/")
def index():
    return render_template("dashboard/overview.html",
                           active_background=_bg_for_template(),
                           active_nav="dashboard",
                           active_sub="overview")


@base_bp.route("/dashboard/overview")
def dashboard_overview():
    return render_template("dashboard/overview.html",
                           active_background=_bg_for_template(),
                           active_nav="dashboard",
                           active_sub="overview")


@base_bp.route("/dashboard/filaments")
def dashboard_filaments():
    return render_template("dashboard/filaments.html",
                           active_background=_bg_for_template(),
                           active_nav="dashboard",
                           active_sub="filaments")


@base_bp.route("/dashboard/logs")
def dashboard_logs():
    return render_template("dashboard/logs.html",
                           active_background=_bg_for_template(),
                           active_nav="dashboard",
                           active_sub="logs")


@base_bp.route("/dashboard/daily")
def dashboard_daily():
    return render_template("dashboard/daily.html",
                           active_background=_bg_for_template(),
                           active_nav="dashboard",
                           active_sub="daily")


@base_bp.route("/dashboard/filament-stats")
def dashboard_filament_stats():
    return render_template("dashboard/filament_stats.html",
                           active_background=_bg_for_template(),
                           active_nav="dashboard",
                           active_sub="filament_stats")


@base_bp.route("/materials")
def materials_page():
    return render_template("materials.html",
                           active_background=_bg_for_template(),
                           active_nav="manage",
                           active_sub="materials")


@base_bp.route("/manufacturers")
def manufacturers_page():
    return render_template("manufacturers.html",
                           active_background=_bg_for_template(),
                           active_nav="manage",
                           active_sub="manufacturers")


@base_bp.route("/settings/general")
def settings_general():
    return render_template("settings/general.html",
                           active_background=_bg_for_template(),
                           active_nav="settings",
                           active_sub="general")


@base_bp.route("/settings/appearance")
def settings_appearance():
    return render_template("settings/appearance.html",
                           active_background=_bg_for_template(),
                           active_nav="settings",
                           active_sub="appearance")


@base_bp.route("/settings/advanced")
def settings_advanced():
    return render_template("settings/advanced.html",
                           active_background=_bg_for_template(),
                           active_nav="settings",
                           active_sub="advanced")


@base_bp.route("/dashboard/brands")
def dashboard_brands():
    return render_template("brand_management.html",
                           active_background=_bg_for_template(),
                           active_nav="manage",
                           active_sub="brands")


@base_bp.route("/roi")
def roi_page():
    return render_template("roi.html",
                           active_background=_bg_for_template(),
                           active_nav="tools",
                           active_sub="roi")


@base_bp.route("/devices")
def devices_page():
    return render_template("device_management.html",
                           active_background=_bg_for_template(),
                           active_nav="devices",
                           active_sub="devices")


@base_bp.route("/dashboard/printer-models")
def printer_models_page():
    return render_template("printer_models.html",
                           active_background=_bg_for_template(),
                           active_nav="devices",
                           active_sub="printer_models")


@base_bp.route("/images")
def images_page():
    return render_template("image_management.html",
                           active_background=_bg_for_template(),
                           active_nav="manage",
                           active_sub="images")


@base_bp.route("/channels")
def channels_page():
    return render_template("channel_management.html",
                           active_background=_bg_for_template(),
                           active_nav="manage",
                           active_sub="channels")


# ─── Uploads File Serving ───


# ─── Settings API ───

@base_bp.route("/api/settings", methods=["GET", "PUT"])
def api_settings():
    data_dir = _data_dir()
    try:
        with get_db(data_dir) as conn:
            if request.method == "GET":
                row = conn.execute("SELECT * FROM settings WHERE id = 1").fetchone()
                result = dict(row)
                card_opacity, card_color, card_blur = _get_appearance_settings(conn)
                result["card_opacity"] = card_opacity
                result["card_color"] = card_color
                result["card_blur"] = card_blur
                lt = conn.execute(
                    "SELECT value FROM system_settings WHERE key = 'low_weight_threshold'"
                ).fetchone()
                result["low_weight_threshold"] = int(lt["value"]) if lt else 100
                return jsonify(result)
            else:
                data = request.get_json() or {}
                threshold = data.get("threshold", 200)
                default_weight = data.get("default_weight", 1000.0)
                conn.execute(
                    "UPDATE settings SET threshold = ?, default_weight = ? WHERE id = 1",
                    (threshold, default_weight),
                )

                if "card_opacity" in data:
                    try:
                        val = float(data["card_opacity"])
                        if 0.0 <= val <= 1.0:
                            conn.execute(
                                "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('card_opacity', ?)",
                                (str(val),),
                            )
                    except (ValueError, TypeError):
                        pass

                if "card_color" in data:
                    color = str(data["card_color"])
                    if re.match(r'^#[0-9a-fA-F]{6}$', color):
                        conn.execute(
                            "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('card_color', ?)",
                            (color,),
                        )

                if "card_blur" in data:
                    try:
                        val = int(data["card_blur"])
                        if 0 <= val <= 30:
                            conn.execute(
                                "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('card_blur', ?)",
                                (str(val),),
                            )
                    except (ValueError, TypeError):
                        pass

                if "low_weight_threshold" in data:
                    try:
                        val = int(data["low_weight_threshold"])
                        if val > 0:
                            conn.execute(
                                "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('low_weight_threshold', ?)",
                                (str(val),),
                            )
                    except (ValueError, TypeError):
                        pass

                conn.commit()
                row = conn.execute("SELECT * FROM settings WHERE id = 1").fetchone()
                result = dict(row)
                card_opacity, card_color, card_blur = _get_appearance_settings(conn)
                result["card_opacity"] = card_opacity
                result["card_color"] = card_color
                result["card_blur"] = card_blur
                lt = conn.execute(
                    "SELECT value FROM system_settings WHERE key = 'low_weight_threshold'"
                ).fetchone()
                result["low_weight_threshold"] = int(lt["value"]) if lt else 100
                return jsonify({"status": "success", "settings": result})
    except Exception as e:
        logger.error("Settings error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


# ─── Background API ───

@base_bp.route("/api/settings/background", methods=["GET"])
def api_background_get():
    data_dir = _data_dir()
    try:
        active = get_active_background(data_dir)
        backgrounds = list_backgrounds(data_dir)
        return jsonify({
            "active": active,
            "backgrounds": backgrounds,
        })
    except Exception as e:
        logger.error("Get background error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


@base_bp.route("/api/settings/background/upload", methods=["POST"])
def api_background_upload():
    data_dir = _data_dir()
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "error": "No file selected"}), 400
        ok, msg, filename = upload_background(request.files["file"], data_dir)
        if not ok:
            return jsonify({"status": "error", "error": msg}), 400
        return jsonify({"status": "success", "message": msg, "filename": filename})
    except Exception as e:
        logger.error("Background upload error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


@base_bp.route("/api/settings/background/set", methods=["POST"])
def api_background_set():
    data_dir = _data_dir()
    try:
        data = request.get_json() or {}
        filename = data.get("filename", "")
        set_active_background(data_dir, filename)
        return jsonify({"status": "success", "active": filename})
    except Exception as e:
        logger.error("Set background error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


@base_bp.route("/api/settings/background/delete", methods=["POST"])
def api_background_delete():
    """Delete a user-uploaded background file. System default (Background.png) is protected."""
    data_dir = _data_dir()
    try:
        data = request.get_json() or {}
        filename = data.get("filename", "")
        if not filename:
            return jsonify({"status": "error", "error": "No filename specified"}), 400
        if filename == "Background.png":
            return jsonify({"status": "error", "error": "System default background cannot be deleted"}), 400

        bg_dir = get_background_dir(data_dir)
        file_path = os.path.join(bg_dir, filename)
        if not os.path.isfile(file_path):
            return jsonify({"status": "error", "error": "File not found"}), 404

        # If this is the active background, reset to default
        active = get_active_background(data_dir)
        if active == filename:
            set_active_background(data_dir, "Background.png")

        os.remove(file_path)
        return jsonify({"status": "success"})
    except Exception as e:
        logger.error("Delete background error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


# ─── Appearance API ───

@base_bp.route("/api/settings/appearance", methods=["GET"])
def api_appearance_get():
    data_dir = _data_dir()
    try:
        with get_db(data_dir) as conn:
            card_opacity, card_color, card_blur = _get_appearance_settings(conn)
            return jsonify({
                "card_opacity": card_opacity,
                "card_color": card_color,
                "card_blur": card_blur,
            })
    except Exception as e:
        logger.error("Appearance get error: %s", e)
        return jsonify({"card_opacity": 0.15, "card_color": "#ffffff", "card_blur": 1})


@base_bp.route("/api/settings/appearance", methods=["PUT"])
def api_appearance_update():
    data_dir = _data_dir()
    try:
        data = request.get_json() or {}
        with get_db(data_dir) as conn:
            if "card_opacity" in data:
                try:
                    val = float(data["card_opacity"])
                    if 0.0 <= val <= 1.0:
                        conn.execute(
                            "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('card_opacity', ?)",
                            (str(val),),
                        )
                except (ValueError, TypeError):
                    pass

            if "card_color" in data:
                color = str(data["card_color"])
                if re.match(r'^#[0-9a-fA-F]{6}$', color):
                    conn.execute(
                        "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('card_color', ?)",
                        (color,),
                    )

            if "card_blur" in data:
                try:
                    val = int(data["card_blur"])
                    if 0 <= val <= 30:
                        conn.execute(
                            "INSERT OR REPLACE INTO system_settings (key, value) VALUES ('card_blur', ?)",
                            (str(val),),
                        )
                except (ValueError, TypeError):
                    pass

            conn.commit()
            card_opacity, card_color, card_blur = _get_appearance_settings(conn)
            return jsonify({
                "status": "success",
                "card_opacity": card_opacity,
                "card_color": card_color,
                "card_blur": card_blur,
            })
    except Exception as e:
        logger.error("Appearance update error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


# ─── Uploads File Serving ───

@base_bp.route("/uploads/backgrounds/<path:filename>")
def serve_background(filename):
    """Serve uploaded background files from /data/uploads/backgrounds/."""
    data_dir = _data_dir()
    bg_dir = get_background_dir(data_dir)
    return send_from_directory(bg_dir, filename)


@base_bp.route("/uploads/filaments/<path:filename>")
def serve_filament_image(filename):
    """Serve uploaded filament images from /data/uploads/filaments/."""
    data_dir = _data_dir()
    img_dir = os.path.join(data_dir, "uploads", "filaments")
    return send_from_directory(img_dir, filename)


# ─── System Config ───

@base_bp.route("/api/system/config", methods=["GET", "POST"])
def api_system_config():
    """Read/write system config key-value pairs."""
    data_dir = _data_dir()
    try:
        with get_db(data_dir) as conn:
            if request.method == "GET":
                rows = conn.execute("SELECT config_key, config_value FROM system_configs").fetchall()
                configs = {}
                for r in rows:
                    try:
                        configs[r["config_key"]] = float(r["config_value"])
                    except (ValueError, TypeError):
                        configs[r["config_key"]] = r["config_value"]
                lang_row = conn.execute("SELECT value FROM system_settings WHERE key='system_language'").fetchone()
                configs["system_language"] = lang_row["value"] if lang_row else "zh"
                return jsonify({"status": "success", "data": configs})
            else:
                data = request.get_json() or {}
                # Handle language separately (string value)
                if "system_language" in data:
                    lang = data.pop("system_language")
                    conn.execute("INSERT OR REPLACE INTO system_settings (key, value) VALUES ('system_language', ?)", (lang,))
                for key, value in data.items():
                    conn.execute(
                        """INSERT INTO system_configs (config_key, config_value)
                           VALUES (?, ?)
                           ON CONFLICT(config_key) DO UPDATE SET config_value=excluded.config_value""",
                        (key, str(value)),
                    )
                conn.commit()
                return jsonify({"status": "success", "message": "Configuration saved successfully"})
    except Exception as e:
        logger.error("System config error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


# ─── System Status ───

@base_bp.route("/api/system/status", methods=["GET"])
def api_system_status():
    """Return system health and version info."""
    data_dir = _data_dir()
    try:
        with get_db(data_dir) as conn:
            row = conn.execute(
                "SELECT value FROM system_settings WHERE key = 'database_version'"
            ).fetchone()
            schema_version = int(row["value"]) if row else 1

        data_ok = os.path.isdir(data_dir) and os.access(data_dir, os.R_OK | os.W_OK)

        return jsonify({
            "program_version": "v0.6.2.3",
            "schema_version": f"Version {schema_version}",
            "schema_latest": LATEST_VERSION >= schema_version,
            "data_status": "Normal (/data read/write ready)" if data_ok else "Error (/data not accessible)",
        })
    except Exception as e:
        logger.error("System status error: %s", e)
        return jsonify({
            "program_version": "v0.6.2.3",
            "schema_version": "Unknown",
            "schema_latest": False,
            "data_status": f"Error: {str(e)}",
        }), 500


# ─── System Backup ───

@base_bp.route("/api/settings/backup", methods=["GET"])
def api_backup():
    """Package /data directory into a zip and stream as download."""
    data_dir = _data_dir()
    try:
        if not os.path.isdir(data_dir):
            return jsonify({"status": "error", "error": "Data directory not found"}), 500

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_name = f"3d_inventory_backup_{timestamp}.zip"

        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(data_dir):
                for filename in files:
                    file_path = os.path.join(root, filename)
                    arcname = os.path.relpath(file_path, data_dir)
                    try:
                        zf.write(file_path, arcname)
                    except (OSError, IOError) as e:
                        logger.warning("Skipping file during backup: %s — %s", file_path, e)

        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype="application/zip",
            as_attachment=True,
            download_name=zip_name,
        )
    except Exception as e:
        logger.error("Backup error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


# ─── Excel Export ───

@base_bp.route("/api/export/excel", methods=["GET"])
def api_export_excel():
    """Export all data tables as a multi-sheet .xlsx workbook."""
    data_dir = _data_dir()
    try:
        wb = Workbook()

        # Sheet 1: filaments
        ws1 = wb.active
        ws1.title = "Filaments"
        headers_f = [
            "id", "name", "material_type", "color", "location",
            "status", "initial_weight", "current_weight", "is_favorite",
            "created_at", "purchase_date", "purchase_price", "Channel", "opened_at",
            "Image Name", "Remark",
        ]
        ws1.append(headers_f)
        with get_db(data_dir) as conn:
            rows = conn.execute("""
                SELECT f.*, fi.name AS image_name, ch.name AS channel_name
                FROM filaments f
                LEFT JOIN filament_images fi ON f.image_id = fi.id
                LEFT JOIN channels ch ON f.channel_id = ch.id
                ORDER BY f.id
            """).fetchall()
            for r in rows:
                ws1.append([
                    r["id"], r["name"], r["material_type"],
                    r["color"], r["location"], r["status"],
                    r["initial_weight"], r["current_weight"], r["is_favorite"],
                    r["created_at"], r["purchase_date"], r["purchase_price"],
                    r["channel_name"] or "", r["opened_at"],
                    r["image_name"] or "", r["remark"] or "",
                ])

            # Sheet 2: materials
            ws2 = wb.create_sheet("Materials")
            ws2.append(["id", "name", "description"])
            for r in conn.execute("SELECT * FROM materials ORDER BY id").fetchall():
                ws2.append([r["id"], r["name"], r["description"]])

            # Sheet 3: usage_records
            ws3 = wb.create_sheet("Usage Records")
            ws3.append(["id", "filament_id", "filament_name", "used_weight", "note", "used_at"])
            usage_rows = conn.execute("""
                SELECT ur.id, ur.filament_id, f.name AS filament_name,
                       ur.used_weight, ur.note, ur.used_at
                FROM usage_records ur
                LEFT JOIN filaments f ON ur.filament_id = f.id
                ORDER BY ur.id
            """).fetchall()
            for r in usage_rows:
                ws3.append([r["id"], r["filament_id"], r["filament_name"],
                           r["used_weight"], r["note"], r["used_at"]])

            # Sheet 4: printers
            ws4 = wb.create_sheet("Printers")
            ws4.append(["id", "name", "model", "created_at"])
            for r in conn.execute("SELECT * FROM printers ORDER BY id").fetchall():
                ws4.append([r["id"], r["name"], r["model"], r["created_at"]])

            # Sheet 5: printer_slots
            ws5 = wb.create_sheet("Printer Slots")
            ws5.append(["id", "printer_id", "printer_name", "slot_name",
                       "current_filament_id", "filament_name", "filament_status"])
            slot_rows = conn.execute("""
                SELECT ps.id, ps.printer_id, p.name AS printer_name, ps.slot_name,
                       ps.current_filament_id, f.name AS filament_name, f.status AS filament_status
                FROM printer_slots ps
                JOIN printers p ON ps.printer_id = p.id
                LEFT JOIN filaments f ON ps.current_filament_id = f.id
                ORDER BY ps.printer_id, ps.id
            """).fetchall()
            for r in slot_rows:
                ws5.append([r["id"], r["printer_id"], r["printer_name"], r["slot_name"],
                           r["current_filament_id"], r["filament_name"], r["filament_status"]])

            # Sheet 6: channels
            ws6 = wb.create_sheet("Channels")
            ws6.append(["id", "name", "description"])
            for r in conn.execute("SELECT * FROM channels ORDER BY id").fetchall():
                ws6.append([r["id"], r["name"], r["description"]])

            # Sheet 7: brands
            ws7 = wb.create_sheet("Brands & Spools")
            ws7.append(["id", "name", "spool_type", "spool_weight", "remark"])
            for r in conn.execute("SELECT * FROM brands ORDER BY name, spool_type").fetchall():
                ws7.append([r["id"], r["name"], r["spool_type"], r["spool_weight"], r["remark"]])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"3d_inventory_export_{timestamp}.xlsx",
        )
    except Exception as e:
        logger.error("Excel export error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


# ─── Backup Restore ───

@base_bp.route("/api/settings/backup/restore", methods=["POST"])
def api_backup_restore():
    """Import and restore a backup ZIP file with hot migration."""
    import tempfile, shutil
    data_dir = _data_dir()
    try:
        if "file" not in request.files:
            return jsonify({"status": "error", "error": "No backup file selected"}), 400
        file = request.files["file"]
        if not file.filename or not file.filename.endswith('.zip'):
            return jsonify({"status": "error", "error": "Please upload a .zip backup file"}), 400

        tmp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(tmp_dir, "backup.zip")
        file.save(zip_path)

        # Extract
        import zipfile
        extract_dir = os.path.join(tmp_dir, "extracted")
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(extract_dir)

        # Find .db file
        db_file = None
        for root, dirs, files in os.walk(extract_dir):
            for f in files:
                if f.endswith('.db'):
                    db_file = os.path.join(root, f)
                    break
        if not db_file:
            shutil.rmtree(tmp_dir)
            return jsonify({"status": "error", "error": "No database file found in backup archive"}), 400

        # Close all DB connections via get_db context manager pattern
        # Copy new DB over
        target_db = os.path.join(data_dir, "database", "filament_inventory.db")
        os.makedirs(os.path.dirname(target_db), exist_ok=True)
        shutil.copy2(db_file, target_db)

        # Restore uploads
        for sub in ['backgrounds', 'filaments']:
            src_uploads = os.path.join(extract_dir, 'uploads', sub)
            if os.path.isdir(src_uploads):
                dst_uploads = os.path.join(data_dir, 'uploads', sub)
                os.makedirs(dst_uploads, exist_ok=True)
                for f in os.listdir(src_uploads):
                    src_f = os.path.join(src_uploads, f)
                    dst_f = os.path.join(dst_uploads, f)
                    if os.path.isfile(src_f):
                        shutil.copy2(src_f, dst_f)

        # Run migration on restored DB
        from modules.db import init_db
        init_db(data_dir)

        shutil.rmtree(tmp_dir)
        return jsonify({"status": "success", "message": "System data restored successfully. Page will refresh."})
    except Exception as e:
        logger.error("Backup restore error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


# ─── Data Migration ───

@base_bp.route("/api/settings/migrate/db", methods=["POST"])
def api_migrate_db():
    """Import data from a legacy filament_inventory.db file."""
    data_dir = _data_dir()
    if "file" not in request.files:
        return jsonify({"status": "error", "error": "No file selected"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"status": "error", "error": "Filename is empty"}), 400

    import tempfile
    tmp_path = None
    try:
        suffix = os.path.splitext(file.filename)[1] or ".db"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        added_f, added_r, updated_s, errors = migrate_from_db(tmp_path, data_dir)
        return jsonify({
            "status": "success",
            "added_filaments": added_f,
            "added_records": added_r,
            "updated_settings": updated_s,
            "errors": errors,
        })
    except Exception as e:
        logger.error("DB migration error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


@base_bp.route("/api/settings/migrate/materials", methods=["POST"])
def api_migrate_materials():
    """Import material types from a legacy materials.txt file."""
    data_dir = _data_dir()
    if "file" not in request.files:
        return jsonify({"status": "error", "error": "No file selected"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"status": "error", "error": "Filename is empty"}), 400

    try:
        added, skipped, errors = migrate_from_txt(file, "materials", data_dir)
        return jsonify({
            "status": "success",
            "added": added,
            "skipped": skipped,
            "errors": errors,
        })
    except Exception as e:
        logger.error("Materials migration error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500


@base_bp.route("/api/settings/migrate/manufacturers", methods=["POST"])
def api_migrate_manufacturers():
    """Import manufacturer brands from a legacy manufacturers.txt file."""
    data_dir = _data_dir()
    if "file" not in request.files:
        return jsonify({"status": "error", "error": "No file selected"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"status": "error", "error": "Filename is empty"}), 400

    try:
        added, skipped, errors = migrate_from_txt(file, "manufacturers", data_dir)
        return jsonify({
            "status": "success",
            "added": added,
            "skipped": skipped,
            "errors": errors,
        })
    except Exception as e:
        logger.error("Manufacturers migration error: %s", e)
        return jsonify({"status": "error", "error": str(e)}), 500
